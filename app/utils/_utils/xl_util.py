import csv
from io import BytesIO

import openpyxl
from flaskz.utils import get_deep, set_deep
from openpyxl import load_workbook, Workbook, utils
from openpyxl.styles import PatternFill


def get_sheet_names(file):
    wb = load_workbook(filename=file)
    return wb.sheetnames


def parse(file, columns, active=False):
    """
    处理Excel文件上传并根据columns参数提取数据

    .. Example::
        #单sheet，返回为list
        data = xl_util.parse(file, [
            {'name': 'Address', 'field': 'address'},
            {'name': 'Name', 'field': 'name'},
            {'name': 'Port', 'field': 'port'},
            {'name': 'Description', 'field': 'description'},
        ])

        #多sheet，返回为dict
        data = xl_util.parse(file, {
            'Login Template': [
                {'name': 'Name', 'field': 'name'},
                {'name': 'Username', 'field': 'username'},
                {'name': 'Password', 'field': 'password'},
                {'name': 'Enable Password', 'field': 'enable_password'},
                {'name': 'Description', 'field': 'description'}
            ],
            'Device': [
                {'name': 'Name', 'field': 'name'},
                {'name': 'Address', 'field': 'ip_address'},
                {'name': 'Manufacturer', 'field': 'manufacturer_name'},
                {'name': 'Software', 'field': 'software_name'},
                {'name': 'Login Template', 'field': 'login_tpl_name'},
                {'name': 'Collection Types', 'field': 'collection_types'}
            ]
        })
    """

    wb = load_workbook(filename=file)
    if type(columns) is dict:  # 多个sheet
        sheet_dict = {}
        for sheet_name, column in columns.items():
            if sheet_name in wb.sheetnames:
                sheet_dict[sheet_name] = _parse_sheet(wb[sheet_name], column)
        return sheet_dict
    else:
        if active is True:  # 单个sheet
            return _parse_sheet(wb.active, columns)

        sheet_dict = {}  # 多个sheet
        for sheet_name in wb.sheetnames:
            sheet_dict[sheet_name] = _parse_sheet(wb[sheet_name], columns)
        return sheet_dict


def generate(data, columns, filename=None):
    """
    将数据导出为 Excel 格式

    .. Example::
        #单个sheet
        file_stream = xl_util.generate(assets_list, [
                {'name': 'Device', 'field': 'device_name'},
                {'name': 'Name', 'field': 'name'},
                {'name': 'PID', 'field': 'pid'},
                {'name': 'VID', 'field': 'vid'},
                {'name': 'SN', 'field': 'sn'},
                {'name': 'Descr', 'field': 'description'},
                {'name': 'Collected At', 'field': 'collected_at'},
            ])

        #多个sheet，相同columns
        file_stream = xl_util.generate(
        {
            'Switch': switch_list,
            'Route': route_list
        },
        [
            {'name': 'Device', 'field': 'device_name'},
            {'name': 'Name', 'field': 'name'},
            {'name': 'SN', 'field': 'sn'},
            {'name': 'Descr', 'field': 'description'}
        ])

        #多个sheet,不同columns
        file_stream = xl_util.generate(
        {
            'Device': device_list,
            'Interface': interface_list
        },
        {
            'Device': device_columns,
            'Interface': interface_columns
        })
    """

    wb = Workbook()
    data_type = type(data)
    columns_type = type(columns)
    if data_type is dict and columns_type is list:  # 多个sheet相同columns
        for sheet_name in data:
            sheet = wb.create_sheet(title=sheet_name)
            fill_sheet_data(sheet, data.get(sheet_name, []), columns)
    elif columns_type is dict and data_type is dict:  # 多个sheet多个columns
        for sheet_name in columns:
            sheet = wb.create_sheet(title=sheet_name)
            fill_sheet_data(sheet, data.get(sheet_name, []), columns.get(sheet_name, []))
    elif columns_type is list and data_type is list:  # 单个sheet
        sheet = wb.active
        fill_sheet_data(sheet, data, columns)
    else:
        return None

    # Workbook()会自动生成名为'Sheet'的空表，删除default sheet
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']

    if filename:
        wb.save(filename)
        return wb

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream


def generate_csv(data, columns, filename, with_header=True):
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        if with_header:
            header_values = [col.get('name', '') for col in columns]
            writer.writerow(header_values)

        rows = []
        for row_data in data:
            rows.append([get_deep(row_data, col.get('field')) for col in columns])
        writer.writerows(rows)


def _parse_sheet(sheet, columns):
    """
    从给定的工作表中提取数据，并根据columns映射进行匹配表头
    """
    _split_merged_cells(sheet)  # 拆分合并单元格

    col_fields = {}
    start_row = int(sheet.dimensions[1])  # 获取表格第一行index

    # 读取第一行&匹配属性
    for index, header in enumerate(sheet[start_row]):
        item_lower = header.value.lower() if (header and header.value) else ''
        if item_lower:
            for col in columns:
                col_name = col.get('name', '')
                col_names = col.get('names', [])
                if col_name:
                    col_names.append(col_name)
                col_names = [s.lower() for s in col_names]
                if item_lower in col_names:
                    col_fields[index] = col.get('field')
                    break
    # 提取数据
    if len(col_fields) == 0:
        return []
    else:
        data_list = []
        for row in sheet.iter_rows(min_row=start_row + 1, values_only=True):
            props = {}
            for col_index in col_fields:
                set_deep(props, col_fields.get(col_index), row[col_index])
                # props[col_fields.get(col_index)] = row[col_index]
            data_list.append(props)
        return data_list


def _split_merged_cells(sheet):
    """拆分合并单元格-->每个单元格"""
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_cell_range in merged_ranges:
        merged_cell_range = str(merged_cell_range)
        start_cell, end_cell = merged_cell_range.split(':')
        value = sheet[start_cell].value
        if value != '':
            sheet.unmerge_cells(merged_cell_range)
            start_column, start_row = utils.cell.coordinate_from_string(start_cell)
            start_row = int(start_row)

            end_column, end_row = utils.cell.coordinate_from_string(end_cell)
            end_row = int(end_row)

            for row in range(start_row, end_row + 1):
                for column in range(utils.column_index_from_string(start_column),
                                    utils.column_index_from_string(end_column) + 1):
                    cell = sheet.cell(row=row, column=column)
                    cell.value = value


def fill_sheet_data(sheet, data, columns, start_row=1, start_col=1):
    """
    将data传入sheet中，并处理每个cell的样式
    """
    for col_index, col in enumerate(columns):
        cell = sheet.cell(row=start_row, column=col_index + start_col, value=col.get('name', ''))
        header_style = col.get('header_style')
        if type(header_style) is dict:
            _apply_cell_styles(cell, header_style)
    # headers = [col.get('name', '') for col in columns]
    # sheet.append(headers)

    for row_index, row_data in enumerate(data):
        # data_row = [row_data.get(col.get('field'), '') for col in columns]
        # data_row = [get_deep(row_data, col.get('field'), default='') for col in columns]
        # sheet.append(data_row)
        for col_index, col in enumerate(columns):
            cell_value = get_deep(row_data, col.get('field'), default='')
            cell = sheet.cell(row=row_index + start_row + 1, column=col_index + start_col, value=cell_value)

            # cell = sheet.cell(row=row_index + 2, column=col_index + 1)
            style = col.get('style')
            if callable(style):
                style = style(cell.value, row_data)
            if type(style) is dict:
                _apply_cell_styles(cell, style)

    for col_index, col in enumerate(columns):
        width = col.get('width')
        dimension_column = _get_dimension_column(col_index + start_col)
        col_width = 0
        if type(width) is int and width > 0:
            col_width = width
            sheet.column_dimensions[dimension_column].width = width
        elif width is not False:
            max_length = 0
            for cell_row_index, cell in enumerate(sheet[dimension_column]):
                if cell_row_index + 1 < start_row:
                    continue
                max_length = max(len(str(cell.value)), max_length)
            col_width = max_length + 2

        if col_width > 0:
            max_width = col.get('max_width')
            if type(max_width) is int and max_width > 0:
                col_width = min(col_width, max_width)

            sheet.column_dimensions[dimension_column].width = col_width


def _get_dimension_column(index):
    column_name = ""
    while index > 0:
        index -= 1  # 转换为0基索引
        column_name = chr(index % 26 + 65) + column_name
        index //= 26
    return column_name


def _apply_cell_styles(cell, style):
    """
    处理sheet中cell的样式
    """

    rgb_map = {
        'red': 'FF0000',
        'orange': 'FFA500',
        'blue': '0000FF',
        'green': '00FF00'
    }

    if 'fill' in style:
        color = rgb_map.get(style['fill'], style['fill'])
        cell.fill = openpyxl.styles.PatternFill(fgColor=color, fill_type='solid')

    if 'alignment' in style:
        alignment_style = style['alignment']
        if type(alignment_style) is str:
            alignment_style = {'horizontal': alignment_style}
        if type(alignment_style) is dict:
            cell.alignment = openpyxl.styles.Alignment(**alignment_style)

    if 'font' in style:
        font_style = style['font']
        if type(font_style) is str:
            color = rgb_map.get(font_style, font_style)
            cell.font = openpyxl.styles.Font(color=color)
        if type(font_style) is dict:
            color = rgb_map.get(font_style.get('color', '000000'), font_style.get('color', '000000'))
            bold = font_style.get('bold', False)
            italic = font_style.get('italic', False)
            cell.font = openpyxl.styles.Font(color=color, bold=bold, italic=italic)


if __name__ == '__main__':
    # 测试parse_xl
    test_file = 'test_excel.xlsx'
    test_columns = [
        {'names': ['address', 'ip', '地址'], 'field': 'address'},
        {'names': ['port', '端口'], 'field': 'port'},
        {'names': ['name', '名称', 'site', '站点'], 'field': 'name'},
        {'names': ['description', '描述', '备注'], 'field': 'description'}
    ]
    # output = parse(test_file, test_columns)
    # print(output)

    # 测试generate_xl
    test_columns = [
        {'field': 'address', 'name': 'Address', 'style': {'fill': 'red', 'alignment': 'center'}, 'header_style': {'fill': 'orange'}},
        {'field': 'port', 'name': 'Port', 'style': lambda value, row_data: {'fill': 'red'} if int(value) > 50 else {'fill': 'green'}},
        {'field': 'name', 'name': 'Name', 'style': {'font': {'color': 'green', 'italic': True, 'bold': 'True'}}},
        {'field': 'description', 'name': 'Description', 'style': {'font': 'red'}}
    ]
    test_data = [
        {'address': '12.0.0.1', 'description': 'Local server', 'name': 'localhost', 'port': '1'},
        {'address': 'www.cisco.com', 'description': 'Cisco site', 'name': 'Cisco', 'port': '8'},
        {'address': 'www.google.com', 'description': 'Google site', 'name': 'Google', 'port': '99'},
        {'address': '10.124.4', 'description': 'Network server', 'name': 'NMS Server', 'port': '80'}
    ]
    workbook = generate(test_data, test_columns, 'test_output.xlsx')
    # Step 2: 获取 BytesIO 对象的内容
    # byte_data = workbook.getvalue()
    #
    # # Step 3: 保存内容到本地文件
    # with open('test_output.xlsx', 'wb') as file:
    #     file.write(byte_data)
    # workbook.save('test_output.xlsx')

    # # 测试apply_styles
    # ws = workbook.active
    # test_cell = ws['A1']
    # test_style = {'fill': 'blue', 'font': {'color': 'red', 'bold': True}}
    # apply_styles(test_cell, test_style)
    # workbook.save('applied_style.xlsx')
